#!/usr/bin/env python3
"""
GMVM v6.1 数据收集脚本 - 优化版
使用并发请求提高性能
"""
import sys
import asyncio
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

# 统一导入模块
import fetch_central_bank_gold_quarterly
import fetch_us_debt_gdp
import fetch_us_bond_tips
import fetch_dxy_index
import fetch_gold_holdings
import fetch_gold_silver_spot
import fetch_aisc
import fetch_gold_technical
import fetch_vix
import fetch_credit_spread
import fetch_gpr
import fetch_crude_oil
import fetch_usa_cpi


async def fetch_task(name, func, *args):
    """封装异步任务"""
    try:
        result = await asyncio.to_thread(func, *args)
        return name, result, None
    except Exception as e:
        return name, None, str(e)


async def collect_data_async():
    """异步收集所有需要的数据"""
    data = {}
    data['收集日期'] = datetime.now().strftime('%Y-%m-%d')

    print("正在并发收集数据...")

    # 创建并发任务
    tasks = [
        # 第一组：网络请求较少的
        fetch_task('cb_gold', fetch_central_bank_gold_quarterly.fetch_central_bank_gold_quarterly),
        fetch_task('debt_gdp', fetch_us_debt_gdp.fetch_us_debt_gdp),
        fetch_task('tips', fetch_us_bond_tips.fetch_bond_data),
        fetch_task('gpr', fetch_gpr.fetch_gpr),
        # 第二组：新浪接口
        fetch_task('dxy', fetch_dxy_index.fetch_dxy_index),
        fetch_task('gold_silver', fetch_gold_silver_spot.fetch_from_sina),
        fetch_task('oil', fetch_crude_oil.fetch_from_sina),
        # 第三组：AkShare接口
        fetch_task('holdings', fetch_gold_holdings.fetch_gold_holdings),
        fetch_task('cpi', fetch_usa_cpi.fetch_usa_cpi),
        # 第四组：其他
        fetch_task('aisc', fetch_aisc.fetch_aisc_data),
        fetch_task('technical', fetch_gold_technical.calculate_gold_technical_analysis),
        fetch_task('vix', fetch_vix.fetch_vix),
        fetch_task('credit_spread', fetch_credit_spread.fetch_credit_spread),
    ]

    # 并发执行所有任务
    results = await asyncio.gather(*tasks)

    # 处理结果
    results_dict = {name: (result, error) for name, result, error in results}

    # 解析数据
    cb_result, _ = results_dict.get('cb_gold', (None, None))
    if cb_result:
        quarters = cb_result.get('quarters', [])
        data['央行购金(吨)'] = sum(q.get('tonnes', 0) for q in quarters)
    else:
        data['央行购金(吨)'] = None

    data['M2同比增速(%)'] = 4.88

    debt_gdp_result, _ = results_dict.get('debt_gdp', (None, None))
    if debt_gdp_result:
        debt = debt_gdp_result.get('national_debt', {}).get('value', 0)
        gdp = debt_gdp_result.get('gdp', {}).get('value', 0)
        if debt and gdp:
            data['美债/GDP比率(%)'] = (debt * 1000000) / (gdp * 1000000000) * 100
        else:
            data['美债/GDP比率(%)'] = None
    else:
        data['美债/GDP比率(%)'] = None

    data['Fed预期差(bp)'] = 2.5

    tips_result, _ = results_dict.get('tips', (None, None))
    if tips_result:
        data['10Y TIPS(%)'] = tips_result.get('data', {}).get('tips_10y', {}).get('latest', {}).get('value')
    else:
        data['10Y TIPS(%)'] = None

    dxy_result, _ = results_dict.get('dxy', (None, None))
    if dxy_result:
        data['DXY'] = dxy_result.get('price')
        data['DXY单日涨幅'] = dxy_result.get('change_percent', 0)
    else:
        data['DXY'] = None
        data['DXY单日涨幅'] = None

    holdings_result, _ = results_dict.get('holdings', (None, None))
    if holdings_result:
        data['ETF持仓4周变化(%)'] = holdings_result.get('latest', {}).get('change_4w_percent')
    else:
        data['ETF持仓4周变化(%)'] = None

    gs_result, _ = results_dict.get('gold_silver', (None, None))
    if gs_result:
        gold_price = gs_result.get('XAU', {}).get('price')
        silver_price = gs_result.get('XAG', {}).get('price')
        if gold_price and silver_price and silver_price > 0:
            data['金银比'] = gold_price / silver_price
        data['金价($)'] = gold_price
        data['金价单日跌幅'] = gs_result.get('XAU', {}).get('change_percent', 0)
    else:
        data['金银比'] = None
        data['金价($)'] = None
        data['金价单日跌幅'] = None

    aisc_result, _ = results_dict.get('aisc', (None, None))
    if aisc_result and data.get('金价($)'):
        aisc_value = aisc_result.get('latest', {}).get('aisc')
        if aisc_value and aisc_value > 0:
            data['金价/AISC'] = data['金价($)'] / aisc_value
    else:
        data['金价/AISC'] = None

    tech_result, _ = results_dict.get('technical', (None, None))
    if tech_result:
        indicators = tech_result.get('indicators', {})
        data['RSI14'] = indicators.get('rsi_14', {}).get('value')
        data['MACD背离'] = indicators.get('macd_12_26_9', {}).get('divergence_score', 50)
        pct_change_5d = indicators.get('ma_slope_20', {}).get('pct_change_5d', 0)
        data['20日均线斜率(%)'] = round(pct_change_5d / 5, 4) if pct_change_5d else None
        data['布林带状态'] = indicators.get('bollinger_20_2', {}).get('score', 50)
    else:
        data['RSI14'] = None
        data['MACD背离'] = None
        data['20日均线斜率(%)'] = None
        data['布林带状态'] = None

    vix_result, _ = results_dict.get('vix', (None, None))
    if vix_result:
        data['VIX'] = vix_result.get('primary_etf', {}).get('price') or 20
    else:
        data['VIX'] = None

    cs_result, _ = results_dict.get('credit_spread', (None, None))
    if cs_result:
        data['信用利差(%)'] = cs_result.get('credit_spread', {}).get('value')
    else:
        data['信用利差(%)'] = None

    gpr_result, _ = results_dict.get('gpr', (None, None))
    if gpr_result:
        data['GPR指数'] = gpr_result.get('gpr')
    else:
        data['GPR指数'] = None

    oil_result, _ = results_dict.get('oil', (None, None))
    if oil_result:
        data['布油价格($)'] = oil_result.get('price')
    else:
        data['布油价格($)'] = None

    cpi_result, _ = results_dict.get('cpi', (None, None))
    if cpi_result:
        data['CPI同比(%)'] = cpi_result.get('current_value')
    else:
        data['CPI同比(%)'] = None

    return data


def fill_excel(data):
    """填写Excel文件"""
    excel_path = Path(__file__).parent / 'GMVM_v6.1_work.xlsx'

    if not excel_path.exists():
        raise FileNotFoundError(f"找不到文件: {excel_path}")

    wb = load_workbook(excel_path)
    ws = wb['数据收集总表']

    headers = [cell.value for cell in ws[1]]
    col_map = {header: idx + 1 for idx, header in enumerate(headers)}

    row_num = 3
    while ws.cell(row=row_num, column=1).value is not None:
        row_num += 1

    for key, value in data.items():
        if key in col_map:
            ws.cell(row=row_num, column=col_map[key], value=value)

    wb.save(excel_path)


def main():
    start_time = datetime.now()
    print("=" * 60)
    print("GMVM v6.1 数据收集脚本 (优化版)")
    print("=" * 60)

    sys.path.insert(0, str(Path(__file__).parent))

    data = asyncio.run(collect_data_async())

    print("\n" + "=" * 60)
    print("收集的数据:")
    print("=" * 60)
    for key, value in data.items():
        print(f"  {key}: {value}")

    fill_excel(data)

    elapsed = (datetime.now() - start_time).total_seconds()

    print(f"\n✅ 数据收集完成！耗时: {elapsed:.2f} 秒")


if __name__ == '__main__':
    main()

