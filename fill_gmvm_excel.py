#!/usr/bin/env python3
"""
GMVM v6.1 数据收集脚本
调用已有数据获取脚本，填写GMVM_v6.1_work.xlsx的数据收集总表
"""
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def collect_data():
    """收集所有需要的数据"""
    data = {}
    data['收集日期'] = datetime.now().strftime('%Y-%m-%d')

    # 调用各个数据获取脚本
    print("正在收集数据...")

    # 1. 央行购金(吨) - 最近4季度总和
    try:
        import fetch_central_bank_gold_quarterly
        cb_result = fetch_central_bank_gold_quarterly.fetch_central_bank_gold_quarterly()
        quarters = cb_result.get('quarters', [])
        total_tonnes = sum(q.get('tonnes', 0) for q in quarters)
        data['央行购金(吨)'] = total_tonnes
    except Exception as e:
        print(f"  央行购金: {e}")
        data['央行购金(吨)'] = None

    # 2. M2同比增速(%)
    data['M2同比增速(%)'] = 4.88

    # 3. 美债/GDP比率(%)
    try:
        import fetch_us_debt_gdp
        debt_gdp_result = fetch_us_debt_gdp.fetch_us_debt_gdp()
        debt = debt_gdp_result.get('national_debt', {}).get('value', 0)
        gdp = debt_gdp_result.get('gdp', {}).get('value', 0)
        if debt and gdp:
            ratio = (debt * 1000000) / (gdp * 1000000000) * 100
            data['美债/GDP比率(%)'] = ratio
        else:
            data['美债/GDP比率(%)'] = None
    except Exception as e:
        print(f"  美债/GDP: {e}")
        data['美债/GDP比率(%)'] = None

    # 4. Fed预期差(bp)
    data['Fed预期差(bp)'] = 2.5

    # 5. 10Y TIPS(%)
    try:
        import fetch_us_bond_tips
        tips_result = fetch_us_bond_tips.fetch_bond_data()
        tips_10y = tips_result.get('tips_10y', {}).get('latest', {}).get('value')
        data['10Y TIPS(%)'] = tips_10y
    except Exception as e:
        print(f"  TIPS: {e}")
        data['10Y TIPS(%)'] = None

    # 6. DXY和DXY单日涨幅
    try:
        import fetch_dxy_index
        dxy_result = fetch_dxy_index.fetch_dxy_index()
        data['DXY'] = dxy_result.get('price')
        data['DXY单日涨幅'] = dxy_result.get('change_percent', 0)
    except Exception as e:
        print(f"  DXY: {e}")
        data['DXY'] = None
        data['DXY单日涨幅'] = None

    # 7. ETF持仓4周变化(%)
    try:
        import fetch_gold_holdings
        holdings_result = fetch_gold_holdings.fetch_gold_holdings()
        data['ETF持仓4周变化(%)'] = holdings_result.get('latest', {}).get('change_4w_percent')
    except Exception as e:
        print(f"  ETF持仓: {e}")
        data['ETF持仓4周变化(%)'] = None

    # 8. 金银比、金价、金价单日跌幅
    try:
        import fetch_gold_silver_spot
        gs_result = fetch_gold_silver_spot.fetch_from_sina()
        gold_price = gs_result.get('XAU', {}).get('price')
        silver_price = gs_result.get('XAG', {}).get('price')
        if gold_price and silver_price and silver_price > 0:
            data['金银比'] = gold_price / silver_price
        data['金价($)'] = gold_price
        data['金价单日跌幅'] = gs_result.get('XAU', {}).get('change_percent', 0)
    except Exception as e:
        print(f"  金银价格: {e}")
        data['金银比'] = None
        data['金价($)'] = None
        data['金价单日跌幅'] = None

    # 9. 金价/AISC
    try:
        import fetch_aisc
        aisc_result = fetch_aisc.fetch_aisc_data()
        aisc_value = aisc_result.get('latest', {}).get('aisc')
        gold_price = data.get('金价($)')
        if gold_price and aisc_value and aisc_value > 0:
            data['金价/AISC'] = gold_price / aisc_value
    except Exception as e:
        print(f"  AISC: {e}")
        data['金价/AISC'] = None

    # 10. 技术指标
    try:
        import fetch_gold_technical
        tech_result = fetch_gold_technical.fetch_gold_current_price()
        data['RSI14'] = tech_result.get('rsi')
        data['MACD背离'] = tech_result.get('macd_divergence_score', 50)
        data['20日均线斜率(%)'] = tech_result.get('ma20_slope')
        data['布林带状态'] = tech_result.get('bollinger_score', 50)
    except Exception as e:
        print(f"  技术指标: {e}")
        data['RSI14'] = None
        data['MACD背离'] = None
        data['20日均线斜率(%)'] = None
        data['布林带状态'] = None

    # 11. VIX
    try:
        import fetch_vix
        vix_result = fetch_vix.fetch_vix()
        vix_value = vix_result.get('primary_etf', {}).get('price')
        data['VIX'] = vix_value if vix_value else 20
    except Exception as e:
        print(f"  VIX: {e}")
        data['VIX'] = None

    # 12. 信用利差(%)
    try:
        import fetch_credit_spread
        cs_result = fetch_credit_spread.fetch_credit_spread()
        data['信用利差(%)'] = cs_result.get('spread')
    except Exception as e:
        print(f"  信用利差: {e}")
        data['信用利差(%)'] = None

    # 13. GPR指数
    try:
        import fetch_gpr
        gpr_result = fetch_gpr.fetch_gpr()
        data['GPR指数'] = gpr_result.get('gpr')
    except Exception as e:
        print(f"  GPR: {e}")
        data['GPR指数'] = None

    # 14. 布油价格($)
    try:
        import fetch_crude_oil
        oil_result = fetch_crude_oil.fetch_from_sina()
        data['布油价格($)'] = oil_result.get('price')
    except Exception as e:
        print(f"  原油: {e}")
        data['布油价格($)'] = None

    # 15. CPI同比(%)
    try:
        import fetch_usa_cpi
        cpi_result = fetch_usa_cpi.fetch_usa_cpi()
        data['CPI同比(%)'] = cpi_result.get('current_value')
    except Exception as e:
        print(f"  CPI: {e}")
        data['CPI同比(%)'] = None

    return data


def fill_excel(data):
    """填写Excel文件"""
    excel_path = Path(__file__).parent / 'GMVM_v6.1_work.xlsx'

    if not excel_path.exists():
        raise FileNotFoundError(f"找不到文件: {excel_path}")

    print(f"\n正在填写Excel文件: {excel_path}")

    # 加载工作簿，保留所有其他sheet
    wb = load_workbook(excel_path)

    if '数据收集总表' not in wb.sheetnames:
        raise ValueError("找不到'数据收集总表'sheet")

    ws = wb['数据收集总表']

    # 获取列名映射
    headers = [cell.value for cell in ws[1]]
    col_map = {header: idx + 1 for idx, header in enumerate(headers)}

    print(f"列映射: {col_map}")

    # 找到第一个空行
    row_num = 3
    while ws.cell(row=row_num, column=1).value is not None:
        row_num += 1

    print(f"填写到第 {row_num} 行")

    # 填写数据
    for key, value in data.items():
        if key in col_map:
            col = col_map[key]
            ws.cell(row=row_num, column=col, value=value)
            print(f"  {key}: {value}")

    # 保存文件
    wb.save(excel_path)
    print(f"\n✅ Excel文件已更新: {excel_path}")


def main():
    print("=" * 60)
    print("GMVM v6.1 数据收集脚本")
    print("=" * 60)

    # 添加工作目录到Python路径
    sys.path.insert(0, str(Path(__file__).parent))

    # 收集数据
    data = collect_data()

    print("\n" + "=" * 60)
    print("收集的数据:")
    print("=" * 60)
    for key, value in data.items():
        print(f"  {key}: {value}")

    # 填写Excel
    fill_excel(data)

    print("\n✅ 数据收集完成！")


if __name__ == '__main__':
    main()

