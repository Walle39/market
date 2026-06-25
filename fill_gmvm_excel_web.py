#!/usr/bin/env python3
"""
GMVM v6.1 数据收集脚本 - 增强版
自动更新Excel表格、网页展示、估值信号，并上传到GitHub
"""
import sys
import asyncio
import subprocess
from datetime import datetime
from pathlib import Path
import re

from openpyxl import load_workbook

# 统一导入模块
import fetch_central_bank_gold_quarterly
import fetch_us_debt_gdp
import fetch_us_bond_tips
import fetch_dxy_index
import fetch_gold_holdings
import fetch_gold_silver_spot
import fetch_aisc
import fetch_gold_technical
import fetch_vix
import fetch_credit_spread
import fetch_gpr
import fetch_crude_oil
import fetch_usa_cpi

# 导入GMVM模型模块
from gmvm_layer1_macro import calculate_s_macro
from gmvm_layer2_verif import calculate_k_verif
from gmvm_layer3_trend import calculate_k_trend
from gmvm_layer4_liquidity import calculate_k_liquidity
from gmvm_layer5_geo import calculate_k_geo


async def fetch_task(name, func, *args):
    """封装异步任务"""
    try:
        result = await asyncio.to_thread(func, *args)
        return name, result, None
    except Exception as e:
        return name, None, str(e)


async def collect_data_async():
    """异步收集所有需要的数据"""
    data = {}
    data['收集日期'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    print("正在并发收集数据...")

    # 创建并发任务
    tasks = [
        fetch_task('cb_gold', fetch_central_bank_gold_quarterly.fetch_central_bank_gold_quarterly),
        fetch_task('debt_gdp', fetch_us_debt_gdp.fetch_us_debt_gdp),
        fetch_task('tips', fetch_us_bond_tips.fetch_bond_data),
        fetch_task('gpr', fetch_gpr.fetch_gpr),
        fetch_task('dxy', fetch_dxy_index.fetch_dxy_index),
        fetch_task('gold_silver', fetch_gold_silver_spot.fetch_from_sina),
        fetch_task('oil', fetch_crude_oil.fetch_from_sina),
        fetch_task('holdings', fetch_gold_holdings.fetch_gold_holdings),
        fetch_task('cpi', fetch_usa_cpi.fetch_usa_cpi),
        fetch_task('aisc', fetch_aisc.fetch_aisc_data),
        fetch_task('technical', fetch_gold_technical.calculate_gold_technical_analysis),
        fetch_task('vix', fetch_vix.fetch_vix),
        fetch_task('credit_spread', fetch_credit_spread.fetch_credit_spread),
    ]

    # 并发执行所有任务
    results = await asyncio.gather(*tasks)

    # 处理结果
    results_dict = {name: (result, error) for name, result, error in results}

    # 解析数据
    cb_result, _ = results_dict.get('cb_gold', (None, None))
    if cb_result:
        quarters = cb_result.get('quarters', [])
        data['央行购金(吨)'] = sum(q.get('tonnes', 0) for q in quarters)
    else:
        data['央行购金(吨)'] = None

    data['M2同比增速(%)'] = 4.88

    debt_gdp_result, _ = results_dict.get('debt_gdp', (None, None))
    if debt_gdp_result:
        debt = debt_gdp_result.get('national_debt', {}).get('value', 0)
        gdp = debt_gdp_result.get('gdp', {}).get('value', 0)
        if debt and gdp:
            data['美债/GDP比率(%)'] = (debt * 1000000) / (gdp * 1000000000) * 100
        else:
            data['美债/GDP比率(%)'] = None
    else:
        data['美债/GDP比率(%)'] = None

    data['Fed预期差(bp)'] = 2.5

    tips_result, _ = results_dict.get('tips', (None, None))
    if tips_result:
        data['10Y TIPS(%)'] = tips_result.get('data', {}).get('tips_10y', {}).get('latest', {}).get('value')
    else:
        data['10Y TIPS(%)'] = None

    dxy_result, _ = results_dict.get('dxy', (None, None))
    if dxy_result:
        data['DXY'] = dxy_result.get('price')
        data['DXY单日涨幅'] = dxy_result.get('change_percent', 0)
    else:
        data['DXY'] = None
        data['DXY单日涨幅'] = None

    holdings_result, _ = results_dict.get('holdings', (None, None))
    if holdings_result:
        data['ETF持仓4周变化(%)'] = holdings_result.get('latest', {}).get('change_4w_percent')
    else:
        data['ETF持仓4周变化(%)'] = None

    gs_result, _ = results_dict.get('gold_silver', (None, None))
    if gs_result:
        gold_price = gs_result.get('XAU', {}).get('price')
        silver_price = gs_result.get('XAG', {}).get('price')
        if gold_price and silver_price and silver_price > 0:
            data['金银比'] = gold_price / silver_price
        data['金价($)'] = gold_price
        data['金价单日跌幅'] = gs_result.get('XAU', {}).get('change_percent', 0)
    else:
        data['金银比'] = None
        data['金价($)'] = None
        data['金价单日跌幅'] = None

    aisc_result, _ = results_dict.get('aisc', (None, None))
    if aisc_result and data.get('金价($)'):
        aisc_value = aisc_result.get('latest', {}).get('aisc')
        if aisc_value and aisc_value > 0:
            data['金价/AISC'] = data['金价($)'] / aisc_value
    else:
        data['金价/AISC'] = None

    tech_result, _ = results_dict.get('technical', (None, None))
    if tech_result:
        indicators = tech_result.get('indicators', {})
        data['RSI14'] = indicators.get('rsi_14', {}).get('value')
        data['MACD背离'] = indicators.get('macd_12_26_9', {}).get('divergence_score', 50)
        pct_change_5d = indicators.get('ma_slope_20', {}).get('pct_change_5d', 0)
        data['20日均线斜率(%)'] = round(pct_change_5d / 5, 4) if pct_change_5d else None
        data['布林带状态'] = indicators.get('bollinger_20_2', {}).get('score', 50)
    else:
        data['RSI14'] = None
        data['MACD背离'] = None
        data['20日均线斜率(%)'] = None
        data['布林带状态'] = None

    vix_result, _ = results_dict.get('vix', (None, None))
    if vix_result:
        data['VIX'] = vix_result.get('primary_etf', {}).get('price') or 20
    else:
        data['VIX'] = None

    cs_result, _ = results_dict.get('credit_spread', (None, None))
    if cs_result:
        data['信用利差(%)'] = cs_result.get('credit_spread', {}).get('value')
    else:
        data['信用利差(%)'] = None

    gpr_result, _ = results_dict.get('gpr', (None, None))
    if gpr_result:
        data['GPR指数'] = gpr_result.get('gpr')
    else:
        data['GPR指数'] = None

    oil_result, _ = results_dict.get('oil', (None, None))
    if oil_result:
        data['布油价格($)'] = oil_result.get('price')
    else:
        data['布油价格($)'] = None

    cpi_result, _ = results_dict.get('cpi', (None, None))
    if cpi_result:
        data['CPI同比(%)'] = cpi_result.get('current_value')
    else:
        data['CPI同比(%)'] = None

    return data


def fill_excel(data):
    """填写Excel文件"""
    excel_path = Path(__file__).parent / 'GMVM_v6.1_work.xlsx'

    if not excel_path.exists():
        raise FileNotFoundError(f"找不到文件: {excel_path}")

    print("正在更新Excel文件...")
    wb = load_workbook(excel_path)
    ws = wb['数据收集总表']

    headers = [cell.value for cell in ws[1]]
    col_map = {header: idx + 1 for idx, header in enumerate(headers)}

    row_num = 3
    while ws.cell(row=row_num, column=1).value is not None:
        row_num += 1

    for key, value in data.items():
        if key in col_map:
            ws.cell(row=row_num, column=col_map[key], value=value)

    wb.save(excel_path)
    print(f"✅ Excel已更新: {excel_path}")


def update_webpage(data):
    """更新网页文件"""
    html_path = Path(__file__).parent / 'index.html'

    if not html_path.exists():
        print("⚠️ 网页文件不存在，跳过网页更新")
        return

    print("正在更新网页...")

    # 读取网页模板
    with open(html_path, 'r', encoding='utf-8') as f:
        html_content = f.read()

    # 定义要替换的数据
    replacements = {
        'timestamp': data.get('收集日期', datetime.now().strftime('%Y-%m-%d')),
        'final-signal': f"{data.get('final_signal', -0.0928):.4f}",
        'signal-grade': data.get('signal_grade', '⚪ 中性'),
        'action': data.get('action', '观望，不做方向性操作'),
        'position': data.get('position', '20%-40%'),
        's_macro_value': f"{data.get('S_macro', -0.10):.2f}",
        'k_verif_value': f"{data.get('K_verif', 0.843):.3f}",
        'k_trend_value': f"{data.get('K_trend', 1.00):.2f}",
        'k_liquidity_value': f"{data.get('K_liquidity', 1.00):.2f}",
        'k_geo_value': f"{data.get('K_geo', 1.10):.2f}",
        '收集日期': data.get('收集日期', ''),
        '央行购金': f"{data.get('央行购金(吨)', 0):.1f}",
        'M2增速': f"{data.get('M2同比增速(%)', 0):.2f}",
        '美债GDP': f"{data.get('美债/GDP比率(%)', 0):.2f}",
        'TIPS': f"{data.get('10Y TIPS(%)', 0):.2f}",
        'DXY': f"{data.get('DXY', 0):.2f}",
        'ETF变化': f"{data.get('ETF持仓4周变化(%)', 0):.2f}",
        '金银比': f"{data.get('金银比', 0):.2f}",
        '金价': f"{data.get('金价($)', 0):,.2f}",
        '金价AISC': f"{data.get('金价/AISC', 0):.2f}",
        'RSI14': f"{data.get('RSI14', 0):.2f}",
        'MACD': f"{data.get('MACD背离', 0):.0f}",
        '均线斜率': f"{data.get('20日均线斜率(%)', 0):.3f}",
        '布林带': f"{data.get('布林带状态', 0):.0f}",
        'VIX': f"{data.get('VIX', 0):.2f}",
        '信用利差': f"{data.get('信用利差(%)', 0):.2f}",
        'GPR': f"{data.get('GPR指数', 0):.2f}",
        '布油': f"{data.get('布油价格($)', 0):.2f}",
        'CPI': f"{data.get('CPI同比(%)', 0):.2f}",
        'footer-timestamp': data.get('收集日期', datetime.now().strftime('%Y-%m-%d')),
    }

    # 替换网页中的数据
    for key, value in replacements.items():
        pattern = f'(<[^>]*id="{key}"[^>]*>)(.*?)(</[^>]+>)'
        html_content = re.sub(pattern, f'\\g<1>{value}\\g<3>', html_content, flags=re.DOTALL)

    # 保存更新后的网页
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

    print(f"✅ 网页已更新: {html_path}")


def calculate_gmvm_signal():
    """计算GMVM v6.1估值信号"""
    print("\n正在计算GMVM v6.1估值信号...")

    try:
        # 计算各层
        s_macro_result = calculate_s_macro()
        k_verif_result = calculate_k_verif()
        k_trend_result = calculate_k_trend()
        k_liquidity_result = calculate_k_liquidity()
        k_geo_result = calculate_k_geo()

        # 计算原始信号
        s_macro = s_macro_result['s_macro']
        k_verif = k_verif_result['k_verif']
        k_trend = k_trend_result['k_trend']
        k_liquidity = k_liquidity_result['k_liquidity']
        k_geo = k_geo_result['k_geo']

        raw_signal = s_macro * k_verif * k_trend * k_liquidity * k_geo

        # 确定信号等级
        if raw_signal >= 0.15:
            signal_grade = "🟢 强烈看多"
            action = "积极做多，分批加仓"
            position = "60%-80%"
        elif raw_signal >= 0.05:
            signal_grade = "🟡 偏多"
            action = "逢低做多，逐步建仓"
            position = "40%-60%"
        elif raw_signal >= -0.05:
            signal_grade = "⚪ 中性"
            action = "观望，不做方向性操作"
            position = "20%-40%"
        elif raw_signal >= -0.15:
            signal_grade = "🟠 偏空"
            action = "逢高减仓，谨慎做空"
            position = "0%-20%"
        else:
            signal_grade = "🔴 强烈看空"
            action = "清仓或做空，严格止损"
            position = "0% 或 空单"

        result = {
            'final_signal': raw_signal,
            'signal_grade': signal_grade,
            'action': action,
            'position': position,
            'S_macro': s_macro,
            'K_verif': k_verif,
            'K_trend': k_trend,
            'K_liquidity': k_liquidity,
            'K_geo': k_geo,
            's_macro_interpretation': s_macro_result.get('signal', ''),
            'k_verif_interpretation': k_verif_result.get('interpretation', ''),
            'k_trend_interpretation': k_trend_result.get('interpretation', ''),
            'k_liquidity_interpretation': k_liquidity_result.get('interpretation', ''),
            'k_geo_interpretation': k_geo_result.get('interpretation', ''),
        }

        print(f"  最终信号: {raw_signal:.4f}")
        print(f"  信号等级: {signal_grade}")
        print(f"  操作建议: {action}")
        print(f"  建议仓位: {position}")

        return result

    except Exception as e:
        print(f"⚠️ GMVM计算失败: {e}")
        return None


def auto_upload_to_github():
    """自动上传到GitHub"""
    print("\n正在自动上传到GitHub...")

    try:
        # 添加所有更改
        result = subprocess.run(
            ['git', 'add', '.'],
            cwd=Path(__file__).parent,
            capture_output=True,
            text=True
        )
        if result.returncode != 0:
            print(f"⚠️ git add 失败: {result.stderr}")
            return False

        # 检查是否有更改
        result = subprocess.run(
            ['git', 'status', '--porcelain'],
            cwd=Path(__file__).parent,
            capture_output=True,
            text=True
        )
        if not result.stdout.strip():
            print("⚠️ 没有需要提交的更改")
            return True

        # 提交
        commit_msg = f"更新数据收集总表({datetime.now().strftime('%Y-%m-%d')})"
        result = subprocess.run(
            ['git', 'commit', '-m', commit_msg],
            cwd=Path(__file__).parent,
            capture_output=True,
            text=True
        )
        if result.returncode != 0:
            print(f"⚠️ git commit 失败: {result.stderr}")
            return False

        # 推送
        result = subprocess.run(
            ['git', 'push'],
            cwd=Path(__file__).parent,
            capture_output=True,
            text=True,
            timeout=60
        )
        if result.returncode != 0:
            print(f"⚠️ git push 失败: {result.stderr}")
            return False

        print("✅ GitHub上传成功！")
        return True

    except subprocess.TimeoutExpired:
        print("⚠️ 上传超时")
        return False
    except Exception as e:
        print(f"⚠️ 上传过程出错: {e}")
        return False


def main():
    start_time = datetime.now()
    print("=" * 60)
    print("GMVM v6.1 数据收集脚本 (增强版)")
    print("自动更新Excel、网页，并上传到GitHub")
    print("=" * 60)

    sys.path.insert(0, str(Path(__file__).parent))

    # 异步收集数据
    data = asyncio.run(collect_data_async())

    # 打印收集的数据
    print("\n" + "=" * 60)
    print("收集的数据:")
    print("=" * 60)
    for key, value in data.items():
        print(f"  {key}: {value}")

    # 计算GMVM估值信号
    gmvm_result = calculate_gmvm_signal()
    if gmvm_result:
        data.update(gmvm_result)

    # 更新Excel
    fill_excel(data)

    # 更新网页
    update_webpage(data)

    # 自动上传到GitHub
    auto_upload_to_github()

    elapsed = (datetime.now() - start_time).total_seconds()

    print(f"\n✅ 全部完成！耗时: {elapsed:.2f} 秒")


if __name__ == '__main__':
    main()

